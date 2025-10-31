import requests # request module
import json #json module
from bson import ObjectId #objectid module
from termcolor import colored #for colors on the console
import re #regex filtering
import os
import datetime #datetime
import time #time
import string #string
import random #random module
from product_model import Product #get the product model
from pymongo import MongoClient, errors #mongodb
from openpyxl import Workbook, load_workbook #for excel sheets
from openai import OpenAI

# product catalog for storing the final products before adding in bulk to db
catalog = []


# AI function to create description for a product
def write_description(title, image_link):
    # Read API key from environment variables
    api_key = os.getenv('OPENAI_KEY') 
    client = OpenAI(api_key=api_key)

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "user",
                "content": f"write 50 words for description of my product for customers to buy, here is the {title} and the link to the image here{image_link}. Don't repeat words and stop using elevate, also do not add the link to the image in description. make the different descriptions unique and catchy to the customer. Do not surround the title with quotes"
            }
        ]
    )

    # print(completion.choices[0].message["content"])
    return completion.choices[0].message.content


# function to store each generated catalog added to db into an excel sheet
def store_new_catalog_data(catalog_list):
    # Prepare the new catalog data
    new_catalog = []

    for item in catalog_list:
        # Create the data dictionary
        new_data = {
            "_id": item['_id'],
            "user": str(item['user']),  # Convert ObjectId to string
            "title": item['title'],
            "images": item['images'],
            "price": item['price'],
            "description": item['description'],
            "category": item['category'],
            "tags": item['tags'],
            "salesOffer": item['salesOffer'],
            "stock": item['stock'],
            "variants": item['variants'],
            "weight": item['weight'],
            "featured": item['featured'],
            "units": item['units'],
            "video": item['video'],
            "productSku": item['productSku'],
            "product_id": item['product_id'],
            "createdAt": str(item['createdAt'].strftime('%Y-%m-%d'))  # Convert datetime to string
        }

        new_catalog.append(new_data)

    # Load existing catalog Excel or create a new one
    try:
        # Open existing catalog.xlsx (if exists)
        wb = load_workbook('catalog.xlsx')
        ws = wb.active
    except FileNotFoundError:
        # If the file does not exist, create a new one
        wb = Workbook()
        ws = wb.active
        ws.append(['IDS', 'USERS', 'TITLES', 'IMAGES', 'PRICES', 'DESCRIPTIONS', 'CATEGORIES', 'TAGS', 'SALESOFFERS', 'STOCKS', 'VARIANTS', 'WEIGHTS', 'FEATUREDS', 'UNITS', 'VIDEOS', 'PRODUCT_SKUS', 'DATES', 'PIDS'])

    # Write new catalog data to the worksheet
    for item in new_catalog:
        ws.append([
            item['_id'],
            item['user'],
            item['title'],
            ','.join(item['images']),
            item['price'],
            item['description'],
            item['category'],
            ','.join(item['tags']),
            item['salesOffer'],
            item['stock'],
            str(item['variants']),
            item['weight'],
            item['featured'],
            item['units'],
            item['video'],
            item['productSku'],
            item['createdAt'],
            item['product_id']
        ])

    # Save the updated catalog to the file
    wb.save('catalog.xlsx')
    print("Catalog data saved to catalog.xlsx")


    
# create array of tags
def create_tags_array(input_string):
    # Check if the string contains '/' or '&'
    if '/' not in input_string and '&' not in input_string:
        return input_string  # Return the original string if neither '/' nor '&' is present
    
    # Otherwise, split the string by '/' and '&'
    result_array = re.split(r'[&/]', input_string)
    
    return result_array


# get the price value if range is involved or not
def split_ranges(range_string):
    # Check if the '-' sign is in the string
    if '-' in range_string:
        # Split the string by the hyphen
        values = range_string.split('-')
        # Return the highest value (second value in the list)
        return float(values[1])
    else:
        # If no hyphen, just return the original float value
        return float(range_string)

# get products from cjdropshipping api
def products_list():
    try:
        url = 'https://developers.cjdropshipping.com/api2.0/v1/product/list' # query url
        headers = {
            # access token of account
            'CJ-Access-Token': "API@CJ3297540@CJ:eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiIxNzA4NiIsInR5cGUiOiJBQ0NFU1NfVE9LRU4iLCJzdWIiOiJicUxvYnFRMGxtTm55UXB4UFdMWnlyK2FrM2QxTU9HTURrWlU4RHF2bU5vWGtFZWlyenRIc29Na3pkamVSQUwzN1pQUEJ6Y1RYWFhEZzNyOHdqYWxveWNtZ0dKMTFMNnJPbENGN245bFRHMXU4VTU4Tkw4dXRCaEZSd25iM1pqRmJLM2l5QkxaamoxeGg1Zkh1MmNNbDdKdEwxMkVjYkNRdnM3cnJsWWhueDIxdWMvMUpPV2pPK0V3ZHk5Tm1QRGF3M0VWMHJ6ZUJhVnpNY2lTWnYySTE0L2JOd2lnMTFXVi9DaVdGVllkWUQ3Q0lXZi8xQjVDRVI2UGh2MnZ6UDR5R3lpRjFXS28yZWpNOUpCRVBEODZsa1ZvaTN6OC91M3M1Y1pjOXZjY0MxMD0iLCJpYXQiOjE3NjE4NjIwNjF9.GJVY63nZ8nwZVdKyJoKgjqMpPDabgG2mP2_T4-95czo",  # Replace `token()` with your token function
        }
        params = { #PARAMS FOR THE REQUESTS 
            'pageNum': 2,
            'pageSize': 95,
            'categoryId': "2FE8A083-5E7B-4179-896D-561EA116F730",  
            'minPrice': 1.0,
            'countryCode': "CN",
            'productType': "SUPPLIER_PRODUCT",
            'verifiedWarehouse': 1
        }
        #  /response feeback
        response = requests.get(url, headers=headers, params=params)
        product_sku_list = [] #skus list of each product
        result = response.json()['data']['list']

        if response.status_code == 200:
            print(colored('--- GET PRODUCTS SUCCESS ---', 'green'))
            for product in result:
                product_sku_list.append(product['productSku'])
            return  product_sku_list
        else:
            print(f"Error: {response.status_code} - {response.text}")
    except Exception as error:
        print(f"Error: {error}")
        
        

# get a single product detail from cjdropshipping api
def get_product_details(product_sku):
    try:
        
        # Generate a random string of the specified length
        # Define the length of the random string
        length = 24
        # Create a pool of characters (small alphabets and digits)
        characters = string.ascii_lowercase + string.digits
        random_string = ''.join(random.choice(characters) for _ in range(length))
        
        url = 'https://developers.cjdropshipping.com/api2.0/v1/product/query'
        headers = {
            'CJ-Access-Token': "API@CJ3297540@CJ:eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiIxNzA4NiIsInR5cGUiOiJBQ0NFU1NfVE9LRU4iLCJzdWIiOiJicUxvYnFRMGxtTm55UXB4UFdMWnlyK2FrM2QxTU9HTURrWlU4RHF2bU5vWGtFZWlyenRIc29Na3pkamVSQUwzN1pQUEJ6Y1RYWFhEZzNyOHdqYWxveWNtZ0dKMTFMNnJPbENGN245bFRHMXU4VTU4Tkw4dXRCaEZSd25iM1pqRmJLM2l5QkxaamoxeGg1Zkh1MmNNbDdKdEwxMkVjYkNRdnM3cnJsWWhueDIxdWMvMUpPV2pPK0V3ZHk5Tm1QRGF3M0VWMHJ6ZUJhVnpNY2lTWnYySTE0L2JOd2lnMTFXVi9DaVdGVllkWUQ3Q0lXZi8xQjVDRVI2UGh2MnZ6UDR5R3lpRjFXS28yZWpNOUpCRVBEODZsa1ZvaTN6OC91M3M1Y1pjOXZjY0MxMD0iLCJpYXQiOjE3NjE4NjIwNjF9.GJVY63nZ8nwZVdKyJoKgjqMpPDabgG2mP2_T4-95czo",  # Replace `token()` with your token function
        }
        params = {
            'productSku': product_sku
        }
        response = requests.get(url, headers=headers, params=params)
        result = response.json()['data']
        
        variants_list = [[variant['variantKey'], variant['variantSku']] for variant in result['variants']]
        
        
        
        if response.status_code == 200:
            if len(result['productImageSet']) > 1:
                print(colored('--- GET PRODUCT DETAILS BY ID SUCCESS ---', 'green'))
                catalog.append({
                        "_id": random_string,
                        "user": ObjectId('676716d67602c42f64f06f94'),
                        "title": result['productNameEn'],
                        "images": result['productImageSet'],
                        "price": round(split_ranges(result['sellPrice']) * 10, 1),
                        "description": write_description(result['productNameEn'], result['productImageSet'][0]),
                        "category": "womens", #category in database
                        "tags": create_tags_array(result['categoryName']),
                        "salesOffer": "25",
                        "stock": result['listedNum'],
                        "variants": variants_list,
                        "weight": split_ranges(result['productWeight']),
                        "featured": "false",
                        "units": 1,
                        "video": "",
                        "productSku": result['productSku'],
                        "product_id": result['pid'],
                        "createdAt": datetime.datetime.now()
                })
                return type(result)
            else:
                print(colored(f"Less ImageSet for {result['productSku']}", 'red'))


        else:
            print(f"Error: {response.status_code} - {response.text}")
    except Exception as error:
        print(f"Error: {error}")


# overall function to import products
def import_products_to_database():
    # get products from api
    products = products_list()
    print(colored("PRODUCTS_SKUS")) 
    print(products)
    
    for product in products:
        print(colored('WAIT FOR 5-SECONDS', 'cyan'))
        time.sleep(5)
        add_to_catalog = get_product_details(product)
        
        
        if add_to_catalog != None:
            print(colored('NEW PRODUCT ENTERED TO CATALOG', 'cyan'))
            
    store_new_catalog_data(catalog)
            
    print(colored(f'CATALOG HAS {len(catalog)} - TOTAL PRODUCTS', 'green'))

    try:
        # Attempt to insert all documents
        new_products = Product.collection.insert_many(catalog)
        
        # Check if any products were successfully inserted
        if new_products.inserted_ids:
            return colored(f"Successfully inserted {len(new_products.inserted_ids)} documents.", 'cyan')
        else:
            return colored("No new products were inserted.", 'cyan')
        
    except errors.BulkWriteError as e:
        # Log the number of failed operations
        print("Some documents were not inserted due to errors:")
        for err in e.details.get('writeErrors', []):
            print(f"Error: {err['errmsg']}, Document: {err['op']}")
        


print(import_products_to_database())