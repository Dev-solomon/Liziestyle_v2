import requests # request module
import json #json module
from bson import ObjectId #objectid module
from termcolor import colored #for colors on the console
import re #regex filtering
import os
import datetime #datetime
import time #time
import string #string
import random #random module
from product_model import Product #get the product model
from review_model import Review # get the review model
from pymongo import MongoClient, errors, DESCENDING #mongodb
from openpyxl import Workbook, load_workbook #for excel sheets
from openai import OpenAI

# product catalog for storing the final products before adding in bulk to db
reviews_catalog = []

# AI function to create description for a product
def create_title(comment):
    # Read API key from environment variables
    api_key = os.getenv('OPENAI_KEY')
    client = OpenAI(api_key=api_key)

    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "user",
                "content": f"create a title  of 5 words in english from this comment: {comment} and do not surround them with quotes."
            }
        ]
    )

    # print(completion.choices[0].message["content"])
    return completion.choices[0].message.content

# function to store each generated catalog added to db into an excel sheet
def store_new_catalog_data(reviews_catalog_list):
    # Prepare the new catalog data
    new_catalog = []

    for item in reviews_catalog_list:
        # Create the data dictionary
        new_data = {
            "name": item['name'],
            "user_id": str(item['user_id']), # Convert ObjectId to string
            "product_id": item['product_id'],
            "title": item['title'],
            "message": item['message'],
            "images": item['images'],
            "rating": int(item['rating']),
            "createdAt": str(item['createdAt'].strftime('%Y-%m-%d'))
        }

        new_catalog.append(new_data)

    # Load existing catalog Excel or create a new one
    try: 
        # Open existing catalog.xlsx (if exists)
        wb = load_workbook('cj/reviews_catalog.xlsx')
        ws = wb.active
    except FileNotFoundError:
        # If the file does not exist, create a new one
        wb = Workbook()
        ws = wb.active
        ws.append(['NAMES', 'USERS', 'PIDS', 'TITLES', 'MESSAGES', 'RATINGS','DATES'])

    # Write new catalog data to the worksheet
    for item in new_catalog:
        ws.append([
            item['name'],
            item['user_id'],
            item['product_id'],
            item['title'],
            item['message'],
            ','.join(item['images']),
            item['rating'],
            item['createdAt']
        ])

    # Save the updated catalog to the file
    wb.save('cj/reviews_catalog.xlsx')
    print("Reviews-Catalog data saved to catalog.xlsx")


    
        
        

# get a single product detail from cjdropshipping api
def get_product_reviews(product_pid):
    try:
        
        
        url = 'https://developers.cjdropshipping.com/api2.0/v1/product/productComments'
        headers = {
            'CJ-Access-Token': "eyJhbGciOiJIUzI1NiJ9.eyJqdGkiOiIxNzA4NiIsInR5cGUiOiJBQ0NFU1NfVE9LRU4iLCJzdWIiOiJicUxvYnFRMGxtTm55UXB4UFdMWnlyK2FrM2QxTU9HTURrWlU4RHF2bU5vWGtFZWlyenRIc29Na3pkamVSQUwzN1pQUEJ6Y1RYWFhEZzNyOHdqYWxvMDBhSjhTQlZlSWJ0TDNIVFc5U1JQQnU4VTU4Tkw4dXRCaEZSd25iM1pqRmJLM2l5QkxaamoxeGg1Zkh1MmNNbDdKdEwxMkVjYkNRdnM3cnJsWWhueDIxdWMvMUpPV2pPK0V3ZHk5Tm1QRGF3M0VWMHJ6ZUJhVnpNY2lTWnYySTE0L2JOd2lnMTFXVi9DaVdGVllkWUQ3Q0lXZi8xQjVDRVI2UGh2MnZ6UDR5R3lpRjFXS28yZWpNOUpCRVBEODZsa1ZvaTN6OC91M3M1Y1pjOXZjY0MxMD0ifQ.YiLZw9VginUjuiA6zTmNORsj8ZKwYhGlPdjglVsR0J4",  # Replace `token()` with your token function
        }
        params = {
            'pid': product_pid,
            "pageNum": 1,
            "pageSize": 35
        }
        response = requests.get(url, headers=headers, params=params)
        result = response.json()['data']['list']
        
        
        product = Product.collection.find_one({"product_id": product_pid})
        
        
        if len(result) > 0:
            print(colored('--- GET PRODUCT REVIEWS BY ID SUCCESS ---', 'green'))
            for review in result:
                reviews_catalog.append({
                        "name": review['commentUser'],
                        "user_id": ObjectId('676716d67602c42f64f06f94'),
                        "product_id": product['_id'],
                        "title": create_title(review['comment']),
                        "message": review['comment'],
                        "images": review['commentUrls'] if review['commentUrls'] is not None else [],
                        "rating": int(review['score']),
                        "createdAt": datetime.datetime.fromisoformat(review['commentDate'])
                })
            return type(result)
        else:
            print(colored(f"NO REVIEWS FOUND FOR PRODUCT WITH ID {product_pid}", 'red'))


    except Exception as error:
        print(f"Error: {error}")


# overall function to import products
def import_reviews_to_database():
    # get products from database
    pageNum = 0
    pageSize = 85
    skip = pageNum * pageSize
    db_products = list(Product.collection.find().sort("createdAt", DESCENDING).limit(pageSize).skip(skip))
    # Extracting only product_id values
    product_ids = [product["product_id"] for product in db_products]
    print(product_ids)
    
    for product_id in product_ids:
        print(colored('WAIT FOR 5-SECONDS', 'cyan'))
        time.sleep(5)
        add_to_catalog = get_product_reviews(product_id)
        
        
        if add_to_catalog != None:
            print(colored('NEW PRODUCT ENTERED TO REVIEW-CATALOG', 'cyan'))
            
    store_new_catalog_data(reviews_catalog)
            
    print(colored(f'REVIEW-CATALOG HAS {len(reviews_catalog)} - TOTAL REVIEWS', 'green'))

    try:
        # Attempt to insert all documents
        new_reviews = Review.collection.insert_many(reviews_catalog)
        
        # Check if any products were successfully inserted
        if new_reviews.inserted_ids:
            return colored(f"Successfully inserted {len(new_reviews.inserted_ids)} documents.", 'cyan')
        else:
            return colored("No new reviews were inserted.", 'cyan')
        
    except errors.BulkWriteError as e:
        # Log the number of failed operations
        print("Some documents were not inserted due to errors:")
        for err in e.details.get('writeErrors', []):
            print(f"Error: {err['errmsg']}, Document: {err['op']}")
        


print(import_reviews_to_database())